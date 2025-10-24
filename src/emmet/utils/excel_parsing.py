"""Excel file parsing utilities."""

from emmet.types import User
from emmet.utils.column_detection import detect_column_by_name
from emmet.utils.column_detection import detect_date_columns
from emmet.utils.column_detection import detect_email_column
from emmet.utils.column_detection import detect_header_row
from emmet.utils.column_detection import detect_name_column
from emmet.utils.name_parsing import parse_name_field
from openpyxl import load_workbook
from typing import Any
from typing import List
from typing import Optional
import datetime
import logging
import uuid


logger = logging.getLogger(__name__)


def should_skip_row(row: Any) -> bool:
    """
    Check if a row should be skipped based on the presence of 'eronnut' (case-insensitive).

    Returns True if the row should be skipped, False otherwise.
    """
    for cell in row:
        if cell.value and isinstance(cell.value, str):
            if "eronnut" in cell.value.lower():
                return True
    return False


def parse_excel_users(
    file_path: str, _column_mapping: Optional[Any] = None
) -> List[User]:
    """
    Parse users from an Excel file using auto-detection heuristics.

    - Finds email column by scanning for email addresses
    - Finds name column by looking for cells with two words (first_name last_name)
    - Generates UUID4 username for each user
    - Skips rows containing 'eronnut' (case-insensitive)
    """
    users: List[User] = []

    try:
        wb = load_workbook(file_path)
        ws = wb.active
        if ws is None:
            logger.warning(f"Worksheet is empty in {file_path}")
            return []

        # Detect which row contains the headers
        header_row_num = detect_header_row(ws)
        header = [
            str(cell.value) if cell.value is not None else ""
            for cell in ws[header_row_num]
        ]

        # Use heuristic approach to detect columns
        email_col_idx = detect_email_column(ws, header, header_row_num)
        if email_col_idx is None:
            logger.error("Could not detect email column in Excel file")
            return []

        name_col_idx = detect_name_column(ws, header, email_col_idx, header_row_num)

        # Hometown is always the next column after name
        hometown_col_idx = (name_col_idx + 1) if name_col_idx is not None else None

        # Detect date columns (skip email, name, and hometown columns)
        skip_cols = [email_col_idx]
        if name_col_idx is not None:
            skip_cols.append(name_col_idx)
        if hometown_col_idx is not None:
            skip_cols.append(hometown_col_idx)

        effective_date_col_idx, expiration_date_col_idx = detect_date_columns(
            ws, header, skip_cols, header_row_num
        )

        # Detect discord and bricklink columns by header name
        discord_col_idx = detect_column_by_name(header, "discord")
        bricklink_col_idx = detect_column_by_name(header, "bricklink")

        logger.info(
            f"Using heuristic parsing: email column at index {email_col_idx}, "
            f"name column at index {name_col_idx}, hometown column at index {hometown_col_idx}, "
            f"effectiveDate column at index {effective_date_col_idx}, "
            f"expirationDate column at index {expiration_date_col_idx}, "
            f"discord column at index {discord_col_idx}, "
            f"bricklink column at index {bricklink_col_idx}"
        )

        # Start processing rows after the header row
        data_start_row = header_row_num + 1
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=data_start_row), start=data_start_row
        ):
            # Skip rows containing 'eronnut'
            if should_skip_row(row):
                logger.info(f"Skipping row {row_idx}: contains 'eronnut'")
                continue

            # Extract email
            email_cell = row[email_col_idx] if email_col_idx < len(row) else None
            email = (
                str(email_cell.value).strip()
                if email_cell and email_cell.value
                else None
            )

            if not email:
                logger.warning(f"Skipping row {row_idx}: no email found")
                continue

            # Extract and parse name
            first_name = None
            last_name = None
            full_name = None
            if name_col_idx is not None and name_col_idx < len(row):
                name_cell = row[name_col_idx]
                if name_cell and name_cell.value:
                    full_name = str(name_cell.value).strip()
                    first_name, last_name = parse_name_field(full_name)

            # Extract hometown (next column after name)
            hometown = None
            if hometown_col_idx is not None and hometown_col_idx < len(row):
                hometown_cell = row[hometown_col_idx]
                if hometown_cell and hometown_cell.value:
                    hometown = str(hometown_cell.value).strip()

            # Extract dates (as strings)
            effective_date = None
            if effective_date_col_idx is not None and effective_date_col_idx < len(row):
                date_cell = row[effective_date_col_idx]
                if date_cell and date_cell.value:
                    # Handle both string and datetime objects
                    if isinstance(date_cell.value, str):
                        effective_date = date_cell.value.strip()
                    elif isinstance(date_cell.value, datetime.datetime):
                        # Convert datetime to dd.mm.yyyy format
                        effective_date = date_cell.value.strftime("%d.%m.%Y")

            expiration_date = None
            if expiration_date_col_idx is not None and expiration_date_col_idx < len(
                row
            ):
                date_cell = row[expiration_date_col_idx]
                if date_cell and date_cell.value:
                    # Handle both string and datetime objects
                    if isinstance(date_cell.value, str):
                        expiration_date = date_cell.value.strip()
                    elif isinstance(date_cell.value, datetime.datetime):
                        # Convert datetime to dd.mm.yyyy format
                        expiration_date = date_cell.value.strftime("%d.%m.%Y")

            # Extract discord
            discord = None
            if discord_col_idx is not None and discord_col_idx < len(row):
                discord_cell = row[discord_col_idx]
                if discord_cell and discord_cell.value:
                    discord = str(discord_cell.value).strip()

            # Extract bricklink
            bricklink = None
            if bricklink_col_idx is not None and bricklink_col_idx < len(row):
                bricklink_cell = row[bricklink_col_idx]
                if bricklink_cell and bricklink_cell.value:
                    bricklink = str(bricklink_cell.value).strip()

            # Generate UUID4 username for new users
            username = str(uuid.uuid4())

            try:
                user = User(
                    username=username,
                    email=email,
                    firstName=first_name,
                    lastName=last_name,
                    fullName=full_name,
                    hometown=hometown,
                    effectiveDate=effective_date,
                    expirationDate=expiration_date,
                    discord=discord,
                    bricklink=bricklink,
                )
                users.append(user)
                logger.info(f"Parsed user from row {row_idx}: {email} -> {username}")
            except Exception as e:
                logger.warning(
                    f"Skipping invalid user data in row {row_idx}: email={email}, "
                    f"firstName={first_name}, lastName={last_name}. Error: {e}"
                )
    except Exception as e:
        logger.error(f"Error reading Excel file {file_path}: {e}")
    return users
