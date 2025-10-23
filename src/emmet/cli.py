from keycloak import KeycloakAdmin
from keycloak.exceptions import KeycloakError
from openpyxl import load_workbook
import click
import logging


# Configure logging
logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")


@click.group()
def main():
    """A CLI for synchronizing user data from an Excel file to Keycloak."""
    pass


@main.command()
@click.argument("excel_file", type=click.Path(exists=True))
@click.option(
    "--keycloak-server",
    envvar="KEYCLOAK_SERVER",
    required=True,
    help="Keycloak server URL.",
)
@click.option(
    "--keycloak-realm",
    envvar="KEYCLOAK_REALM",
    required=True,
    help="Keycloak realm name.",
)
@click.option(
    "--keycloak-client-id",
    envvar="KEYCLOAK_CLIENT_ID",
    required=True,
    help="Keycloak client ID.",
)
@click.option(
    "--keycloak-client-secret",
    envvar="KEYCLOAK_CLIENT_SECRET",
    required=True,
    help="Keycloak client secret.",
)
@click.option(
    "--dry-run", is_flag=True, help="Only print actions, do not execute them."
)
def sync(
    excel_file,
    keycloak_server,
    keycloak_realm,
    keycloak_client_id,
    keycloak_client_secret,
    dry_run,
):
    """Synchronize users from an Excel file to Keycloak."""
    click.echo(f"Synchronizing users from {excel_file}...")

    # Read users from Excel file
    excel_users = []
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2):
            user_data = {header[i]: cell.value for i, cell in enumerate(row)}
            excel_users.append(user_data)
    except Exception as e:
        logging.error(f"Error reading Excel file: {e}")
        return

    # Connect to Keycloak
    try:
        keycloak_admin = KeycloakAdmin(
            server_url=keycloak_server,
            client_id=keycloak_client_id,
            client_secret_key=keycloak_client_secret,
            realm_name=keycloak_realm,
            user_realm_name="master",
        )
        keycloak_admin.connection.get_token()
    except KeycloakError as e:
        logging.error(f"Error connecting to Keycloak: {e}")
        return

    # Get all users from Keycloak
    try:
        keycloak_users = keycloak_admin.get_users({})
        keycloak_usernames = [user["username"] for user in keycloak_users]
    except KeycloakError as e:
        logging.error(f"Error getting users from Keycloak: {e}")
        return

    # Sync users from Excel to Keycloak
    for user in excel_users:
        username = user.get("username")
        if not username:
            logging.warning(f"Skipping row with missing username: {user}")
            continue

        try:
            user_id = keycloak_admin.get_user_id(username)
            if user_id:
                logging.info(f"Updating user {username}...")
                if not dry_run:
                    keycloak_admin.update_user(
                        user_id,
                        {
                            "email": user.get("email"),
                            "firstName": user.get("firstName"),
                            "lastName": user.get("lastName"),
                        },
                    )
            else:
                logging.info(f"Creating user {username}...")
                if not dry_run:
                    keycloak_admin.create_user(
                        {
                            "username": username,
                            "email": user.get("email"),
                            "firstName": user.get("firstName"),
                            "lastName": user.get("lastName"),
                            "enabled": True,
                        }
                    )
        except KeycloakError as e:
            logging.error(f"Error syncing user {username}: {e}")

    # Disable users in Keycloak that are not in the Excel file
    excel_usernames = [user.get("username") for user in excel_users]
    for username in keycloak_usernames:
        if username not in excel_usernames:
            logging.info(f"Disabling user {username}...")
            if not dry_run:
                try:
                    user_id = keycloak_admin.get_user_id(username)
                    keycloak_admin.update_user(user_id, {"enabled": False})
                except KeycloakError as e:
                    logging.error(f"Error disabling user {username}: {e}")


if __name__ == "__main__":
    main()
