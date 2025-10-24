from emmet.types import User
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import Any
from typing import Dict
from typing import List
from typing import Optional
from typing import Tuple
import datetime
import logging
import re
import uuid


logger = logging.getLogger(__name__)


def detect_email_column(
    ws: Worksheet, header: List[str], header_row_num: int = 1
) -> Optional[int]:
    """
    Detect which column contains email addresses by scanning the data.

    Returns the column index (0-based) or None if no email column is found.
    """
    email_pattern = re.compile(r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$")

    # Count email matches per column
    column_email_counts: Dict[int, int] = {}

    # Start scanning after the header row
    data_start_row = header_row_num + 1
    for row in ws.iter_rows(
        min_row=data_start_row,
        max_row=min(data_start_row + 19, ws.max_row or data_start_row),
    ):  # Sample up to 20 data rows
        for col_idx, cell in enumerate(row):
            if cell.value and isinstance(cell.value, str):
                if email_pattern.match(cell.value.strip()):
                    column_email_counts[col_idx] = (
                        column_email_counts.get(col_idx, 0) + 1
                    )

    # Return column with most email matches (at least 1)
    if column_email_counts:
        email_col_idx = max(column_email_counts, key=lambda k: column_email_counts[k])
        logger.info(
            f"Detected email column: '{header[email_col_idx]}' (index {email_col_idx})"
        )
        return email_col_idx

    return None


def detect_name_column(
    ws: Worksheet,
    header: List[str],
    email_col_idx: Optional[int],
    header_row_num: int = 1,
) -> Optional[int]:
    """
    Detect which column contains full names (two or more words: first_name [middle_name(s)] last_name).

    Returns the column index (0-based) or None if no name column is found.
    """
    two_or_more_word_pattern = re.compile(r"^\s*\S+\s+\S+.*$")

    # Count two-or-more-word matches per column
    column_name_counts: Dict[int, int] = {}

    # Start scanning after the header row
    data_start_row = header_row_num + 1
    for row in ws.iter_rows(
        min_row=data_start_row,
        max_row=min(data_start_row + 19, ws.max_row or data_start_row),
    ):  # Sample up to 20 data rows
        for col_idx, cell in enumerate(row):
            # Skip the email column
            if col_idx == email_col_idx:
                continue

            if cell.value and isinstance(cell.value, str):
                if two_or_more_word_pattern.match(cell.value.strip()):
                    column_name_counts[col_idx] = column_name_counts.get(col_idx, 0) + 1

    # Return column with most two-or-more-word matches (at least 1)
    if column_name_counts:
        name_col_idx = max(column_name_counts, key=lambda k: column_name_counts[k])
        logger.info(
            f"Detected name column: '{header[name_col_idx]}' (index {name_col_idx})"
        )
        return name_col_idx

    return None


def detect_date_columns(
    ws: Worksheet,
    header: List[str],
    skip_col_indices: List[int],
    header_row_num: int = 1,
) -> Tuple[Optional[int], Optional[int]]:
    """
    Detect which columns contain dates in dd.mm.yyyy format.

    Returns a tuple of (first_date_col_idx, second_date_col_idx).
    The first date is effectiveDate, the second is expirationDate.
    """
    # Pattern for dd.mm.yyyy format (as string or datetime)
    date_pattern = re.compile(r"^\d{1,2}\.\d{1,2}\.\d{4}$")

    # Track which columns have date-like values
    column_date_counts: Dict[int, int] = {}

    # Start scanning after the header row
    data_start_row = header_row_num + 1
    for row in ws.iter_rows(
        min_row=data_start_row,
        max_row=min(data_start_row + 19, ws.max_row or data_start_row),
    ):  # Sample up to 20 data rows
        for col_idx, cell in enumerate(row):
            # Skip already identified columns
            if col_idx in skip_col_indices:
                continue

            if cell.value:
                # Check if it's a string matching date pattern
                if isinstance(cell.value, str) and date_pattern.match(
                    cell.value.strip()
                ):
                    column_date_counts[col_idx] = column_date_counts.get(col_idx, 0) + 1
                # Check if it's a datetime object from Excel
                elif hasattr(cell.value, "strftime"):
                    column_date_counts[col_idx] = column_date_counts.get(col_idx, 0) + 1

    # Get columns sorted by index that have date values
    date_columns = sorted(
        [col_idx for col_idx, count in column_date_counts.items() if count > 0]
    )

    effective_date_col = date_columns[0] if len(date_columns) >= 1 else None
    expiration_date_col = date_columns[1] if len(date_columns) >= 2 else None

    if effective_date_col is not None:
        logger.info(
            f"Detected effectiveDate column: '{header[effective_date_col]}' (index {effective_date_col})"
        )
    if expiration_date_col is not None:
        logger.info(
            f"Detected expirationDate column: '{header[expiration_date_col]}' (index {expiration_date_col})"
        )

    return effective_date_col, expiration_date_col


def detect_header_row(ws: Worksheet) -> int:
    """
    Detect which row contains the header by looking for "bricklink" (case-insensitive).

    Returns the row number (1-based) of the header row.
    """
    # Check first 10 rows to find the header
    for row_num in range(1, min(11, (ws.max_row or 1) + 1)):
        row_values = [
            str(cell.value) if cell.value is not None else "" for cell in ws[row_num]
        ]

        # Check if any cell in this row contains "bricklink" (case-insensitive)
        for cell_value in row_values:
            if cell_value and "bricklink" in cell_value.lower():
                logger.info(
                    f"Detected header row at row {row_num} (contains 'bricklink')"
                )
                return row_num

    # Default to row 1 if no header detected
    logger.info("Using row 1 as header row (default)")
    return 1


def detect_column_by_name(header: List[str], search_term: str) -> Optional[int]:
    """
    Detect a column by searching for a case-insensitive term in the header.

    Returns the column index (0-based) or None if not found.
    """
    # if search_term == "discord":
    #    import pdb; pdb.set_trace()
    search_term_lower = search_term.lower()
    logger.info(f"Searching for column containing '{search_term}'...")
    for col_idx, header_value in enumerate(header):
        if header_value:
            logger.debug(f"  Checking column {col_idx}: '{header_value}'")
            if search_term_lower in header_value.lower():
                logger.info(
                    f"Detected {search_term} column: '{header[col_idx]}' (index {col_idx})"
                )
                return col_idx
    logger.warning(f"No column found containing '{search_term}'")
    return None


def should_skip_row(row: Any) -> bool:
    """
    Check if a row should be skipped based on the presence of 'eronnut' (case-insensitive).

    Returns True if the row should be skipped, False otherwise.
    """
    for cell in row:
        if cell.value and isinstance(cell.value, str):
            if "eronnut" in cell.value.lower():
                return True
    return False


def parse_name_field(name_str: str) -> Tuple[Optional[str], Optional[str]]:
    """
    Parse a name string into first_name and last_name.

    - If the name has two or more parts: first part is first_name, last part is last_name
    - Middle parts are ignored
    - If only one part: it becomes first_name, last_name is None

    Returns (first_name, last_name) tuple.
    """
    if not name_str:
        return None, None

    parts = name_str.strip().split()  # Split on whitespace
    if len(parts) >= 2:
        return parts[0], parts[-1]  # First and last part
    elif len(parts) == 1:
        return parts[0], None

    return None, None


def parse_excel_users(
    file_path: str, _column_mapping: Optional[Any] = None
) -> List[User]:
    """
    Parse users from an Excel file using auto-detection heuristics.

    - Finds email column by scanning for email addresses
    - Finds name column by looking for cells with two words (first_name last_name)
    - Generates UUID4 username for each user
    - Skips rows containing 'eronnut' (case-insensitive)
    """
    users: List[User] = []

    try:
        wb = load_workbook(file_path)
        ws = wb.active
        if ws is None:
            logger.warning(f"Worksheet is empty in {file_path}")
            return []

        # Detect which row contains the headers
        header_row_num = detect_header_row(ws)
        header = [
            str(cell.value) if cell.value is not None else ""
            for cell in ws[header_row_num]
        ]

        # Use heuristic approach to detect columns
        email_col_idx = detect_email_column(ws, header, header_row_num)
        if email_col_idx is None:
            logger.error("Could not detect email column in Excel file")
            return []

        name_col_idx = detect_name_column(ws, header, email_col_idx, header_row_num)

        # Hometown is always the next column after name
        hometown_col_idx = (name_col_idx + 1) if name_col_idx is not None else None

        # Detect date columns (skip email, name, and hometown columns)
        skip_cols = [email_col_idx]
        if name_col_idx is not None:
            skip_cols.append(name_col_idx)
        if hometown_col_idx is not None:
            skip_cols.append(hometown_col_idx)

        effective_date_col_idx, expiration_date_col_idx = detect_date_columns(
            ws, header, skip_cols, header_row_num
        )

        # Detect discord and bricklink columns by header name
        discord_col_idx = detect_column_by_name(header, "discord")
        bricklink_col_idx = detect_column_by_name(header, "bricklink")

        logger.info(
            f"Using heuristic parsing: email column at index {email_col_idx}, "
            f"name column at index {name_col_idx}, hometown column at index {hometown_col_idx}, "
            f"effectiveDate column at index {effective_date_col_idx}, "
            f"expirationDate column at index {expiration_date_col_idx}, "
            f"discord column at index {discord_col_idx}, "
            f"bricklink column at index {bricklink_col_idx}"
        )

        # Start processing rows after the header row
        data_start_row = header_row_num + 1
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=data_start_row), start=data_start_row
        ):
            # Skip rows containing 'eronnut'
            if should_skip_row(row):
                logger.info(f"Skipping row {row_idx}: contains 'eronnut'")
                continue

            # Extract email
            email_cell = row[email_col_idx] if email_col_idx < len(row) else None
            email = (
                str(email_cell.value).strip()
                if email_cell and email_cell.value
                else None
            )

            if not email:
                logger.warning(f"Skipping row {row_idx}: no email found")
                continue

            # Extract and parse name
            first_name = None
            last_name = None
            full_name = None
            if name_col_idx is not None and name_col_idx < len(row):
                name_cell = row[name_col_idx]
                if name_cell and name_cell.value:
                    full_name = str(name_cell.value).strip()
                    first_name, last_name = parse_name_field(full_name)

            # Extract hometown (next column after name)
            hometown = None
            if hometown_col_idx is not None and hometown_col_idx < len(row):
                hometown_cell = row[hometown_col_idx]
                if hometown_cell and hometown_cell.value:
                    hometown = str(hometown_cell.value).strip()

            # Extract dates (as strings)
            effective_date = None
            if effective_date_col_idx is not None and effective_date_col_idx < len(row):
                date_cell = row[effective_date_col_idx]
                if date_cell and date_cell.value:
                    # Handle both string and datetime objects
                    if isinstance(date_cell.value, str):
                        effective_date = date_cell.value.strip()
                    elif isinstance(date_cell.value, datetime.datetime):
                        # Convert datetime to dd.mm.yyyy format
                        effective_date = date_cell.value.strftime("%d.%m.%Y")

            expiration_date = None
            if expiration_date_col_idx is not None and expiration_date_col_idx < len(
                row
            ):
                date_cell = row[expiration_date_col_idx]
                if date_cell and date_cell.value:
                    # Handle both string and datetime objects
                    if isinstance(date_cell.value, str):
                        expiration_date = date_cell.value.strip()
                    elif isinstance(date_cell.value, datetime.datetime):
                        # Convert datetime to dd.mm.yyyy format
                        expiration_date = date_cell.value.strftime("%d.%m.%Y")

            # Extract discord
            discord = None
            if discord_col_idx is not None and discord_col_idx < len(row):
                discord_cell = row[discord_col_idx]
                if discord_cell and discord_cell.value:
                    discord = str(discord_cell.value).strip()

            # Extract bricklink
            bricklink = None
            if bricklink_col_idx is not None and bricklink_col_idx < len(row):
                bricklink_cell = row[bricklink_col_idx]
                if bricklink_cell and bricklink_cell.value:
                    bricklink = str(bricklink_cell.value).strip()

            # Generate UUID4 username for new users
            username = str(uuid.uuid4())

            try:
                user = User(
                    username=username,
                    email=email,
                    firstName=first_name,
                    lastName=last_name,
                    fullName=full_name,
                    hometown=hometown,
                    effectiveDate=effective_date,
                    expirationDate=expiration_date,
                    discord=discord,
                    bricklink=bricklink,
                )
                users.append(user)
                logger.info(f"Parsed user from row {row_idx}: {email} -> {username}")
            except Exception as e:
                logger.warning(
                    f"Skipping invalid user data in row {row_idx}: email={email}, "
                    f"firstName={first_name}, lastName={last_name}. Error: {e}"
                )
    except Exception as e:
        logger.error(f"Error reading Excel file {file_path}: {e}")
    return users
